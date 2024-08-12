import os
import datetime
from openpyxl import load_workbook
import dirs


def importThresholds():
    threshold_dict = {"thresholds":[]}
    logfile = load_workbook(dirs.LOG_FILE)
    activesheet = logfile["Thresholds"]
    for i in range(4, 8):
        if activesheet.cell(i,2).value:
            tempt = {}
            tempt["gas"] = activesheet.cell(i,2).value
            tempt["name"] = activesheet.cell(i,3).value
            tempt["concentration"] = activesheet.cell(i,4).value
            threshold_dict["thresholds"].append(tempt)
    return threshold_dict
            
