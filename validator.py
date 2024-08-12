from openpyxl import load_workbook
from IPython.display import Markdown as md
import pandas as pd

#paramater_isvalid in format {'CO(ppm)': True, ...}
#use format 8859.1 format from file


class DetectorValidation():
    def __init__(self, serial, start, stop=None, reason=None):
        self.serial = serial
        self.start = pd.to_datetime(start)
        self.stop = None
        if stop:
            self.stop = pd.to_datetime(stop)
        self.reason = reason
        self.parameters = {}
    def addParameter(self, key, val):
        self.parameters[key] = val

    def isParameterValid(self, key):
        return self.parameters[key]

    def __repr__(self):
        return f"<Detector Validation serial={self.serial} start={self.start} stop={self.stop}>"
        
def importValidations():       
    detector_validation_dict = {"validations":[]}
    logfile = load_workbook(dirs.LOG_FILE)
    activesheet = logfile["Validations"]
    for i in range(5, 100):
        if activesheet.cell(i,2).value:
            detector_validation_dict["validations"].append(
                                                            {
                                                                "serial": activesheet.cell(i,2).value,
                                                                "start": activesheet.cell(i,3).value,
                                                                "stop": activesheet.cell(i,4).value,
                                                                "reason": activesheet.cell(i,5).value,
                                                                "params": {}
                                                            }
                                                        )
            #set validation column to first column of parameters in file
            header_column = 4
            validation_columns = range(6, 11)
            dv_dict = detector_validation_dict["validations"][-1]["params"]
            for col in validation_columns:
                if activesheet.cell(i, col).value == "Invalid":
                    dv_dict[activesheet.cell(header_column, col).value] = False
                else:
                    dv_dict[activesheet.cell(header_column, col).value] = True
    return detector_validation_dict        

def printValidationDict(detector_validation_dict):
    for validation in detector_validation_dict["validations"]:
        print(f"Found <Validation serial={validation['serial']} start={validation['start']} stop={validation['stop']} reason={str(validation['reason'])[:10]}...>")
        for param in validation["params"].keys():
            print(f"\t<{param} isValid={validation['params'][param]}>") 

#sets the validation flag to 0 for the time and device in the detector validation
def applyValidations(location_df_list, detector_validation_dict):
    #apply validation
    print("Applying validations...")
    for location_df in location_df_list:
        if len(location_df) > 0: 
            for dv in detector_validation_dict['validations']:
                #print(locdf.loc[locdf["Serial"] == dv.serial, locdf.index > dv.from_datetime])
                #set stop to the future if None
                stop = pd.to_datetime("2119-08-22")
                if dv['stop']:
                    stop = dv['stop']
                for param in dv['params'].keys():
                    if not dv['params'][param]:
                        location_df.loc[(location_df.index > dv['start']) & (location_df.index <= stop) & (location_df['Serial'] == dv['serial']), param + "_Validation"] = 0 
                        ninvalidated = len(location_df.loc[(location_df.index > dv['start']) & (location_df.index <= stop) & (location_df['Serial'] == dv['serial']), param + "_Validation"])
                        if ninvalidated > 0:
                            print(f"Invalidated {len(location_df.loc[(location_df.index > dv['start']) & (location_df.index <= stop) & (location_df['Serial'] == dv['serial']), param + '_Validation'])} readings at Location {location_df.label}")
    print("Finished applying validations.")
    
                    

        