from openpyxl import load_workbook
from IPython.display import Markdown as md
import pandas as pd
import math
LOG_FILE = "./Data/LocationLog/Location Log.xlsx"
     
def importCorrections():
    correction_dict = {"corrections":{}}
    logfile = load_workbook(LOG_FILE)
    activesheet = logfile["Corrections"]
    for i in range(4, 8):
        if activesheet.cell(i,2).value:
            correction_dict["corrections"][activesheet.cell(i,2).value] = activesheet.cell(i,3).value
    return correction_dict
            
def applyCorrections(location_df_list, correction_dict):
    print("Applying corrections...")
    for location_df in location_df_list:
        if len(location_df) > 0: 
            for param in correction_dict["corrections"].keys():
                #only bother with corrections not equal to 1.0
                if not math.isclose(correction_dict["corrections"][param], 1.00):
                    print(f"Correcting {param}. Multiplying column by {correction_dict['corrections'][param]}")
                    location_df.loc[:, param] *= correction_dict["corrections"][param]
    print("Finished applying corrections")