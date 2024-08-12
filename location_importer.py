import os
import datetime
from openpyxl import load_workbook
import pandas as pd
import dirs


    
#returns a list of locations each with a sublist of detectors and the times they were there
def importLocations():
    location_dict = {"locations":[]}
    if os.path.exists(dirs.LOG_FILE):
        logfile = load_workbook(dirs.LOG_FILE)
        for sheet in logfile.worksheets:
            if sheet.title[0:8] == "Location":
                label = sheet.cell(2,2).value[-1]
                description = sheet.cell(4,3).value
                lat = sheet.cell(6,3).value
                if sheet.cell(6,3).value:
                    latitude = sheet.cell(6,3).value.split(",")[0].strip()
                    longitude = sheet.cell(6,3).value.split(",")[1].strip()
                else:
                    latitude = None
                    longitude = None
                location_dict["locations"].append(
                                                    {
                                                        "label": label,
                                                        "description": description,
                                                        "latitude": latitude,
                                                        "longitude": longitude,
                                                        "detectors": []
                                                    }
                                                )

                                         
                for row in range(9, 50):
                    if sheet.cell(row, 2).value is not None:
                        location_dict["locations"][-1]["detectors"].append(
                                                                                {
                                                                                    "serial": sheet.cell(row, 2).value,
                                                                                    "start": sheet.cell(row, 3).value,
                                                                                    "stop": sheet.cell(row, 4).value
                                                                                }
                                                                           )
    return location_dict
    
 
def checkLocationDict(location_dict):
    #validate location log
    error_log = []
    for location in location_dict["locations"]:
        #set last_datetime to past to ensure the next start datetime is greater than the last
        last_start_datetime = datetime.datetime(2000, 1, 1)
        lastdetectorindex = len(location["detectors"]) - 1
        for ind, det in enumerate(location["detectors"]):
            if not det['stop']:
                #check if detector was left unstopped in the middle of the log. Should only be the last one
                if ind != lastdetectorindex:
                    error_log.append(f"Location {location['label']} has unstopped detector before end of list")
                else:
                    #check start is after the last start
                    if pd.to_datetime(det['start']) < last_start_datetime:
                        error_log.append(f"Location {location['label']} has a start before the previous stop")
            else:
                #check
                if det['start'] >= det['stop']:
                     error_log.append(f"Location {location['label']} has a start before a stop")
                else:
                    last_start_datetime = pd.to_datetime(det['stop'])
    if len(error_log) == 0:
        print("No errors found in location list. Proceed to next step")
        return False
    else:
        print("Errors found in location log. Please fix before proceeding.")
        for e in error_log:
            print(f"\t{e}")
        True
    
def printLocationDict(location_dict):
    for location in location_dict["locations"]:
        print(f"Found <Location label={location['label']} description={str(location['description'])[:10]}... lat.={location['latitude']} long.={location['longitude']}>")
        for detector in location["detectors"]:
            print(f"\t<Detector serial={detector['serial']} start={detector['start']} stop={detector['stop']}>") 
    
    
if __name__ == "__main__":
    location_list = importLocations()
    for loc in location_list:
        printLocation(loc)
    validateLocationList(location_list)